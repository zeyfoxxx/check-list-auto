import os
import win32com.client
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from datetime import datetime

# Chemin d'accès fichier orgine/temp/dest
TEMPLATE_CHEMIN = r"C:\Users\lbenadyext\Desktop\check-list auto\CD13_MQ_.xlsx"
FICHIER_SORTIE = r"C:\Users\lbenadyext\Desktop\check-list auto"
FICHIER_TEMP = r"C:\Users\lbenadyext\Desktop\check-list auto\check_temp" 


# Au cas ou le fichier temp n'est pas créer pour le créer seul
os.makedirs(FICHIER_TEMP, exist_ok=True)
os.makedirs(FICHIER_SORTIE, exist_ok=True)

#création de la fonction date parce que flemme de faire moi mêm tout les jours
date_aujourdhui = datetime.now().strftime("%d-%m-%Y")
nouveau_fichier = os.path.join(FICHIER_SORTIE, f"CD13_MQ_{date_aujourdhui}.xlsx")

def recuperer_mails_et_pieces():
    print("Connexion à Outlook...")
    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        shared_mailbox = outlook.Folders.Item("SESN Wan")
        inbox = shared_mailbox.Folders("Boîte de réception") 
        folder_free_pro = inbox.Folders("Free Pro")
        target_folder = folder_free_pro.Folders("Rapports du matin")
        print(f"Succès : Dossier '{target_folder.Name}' trouvé.")
        
        messages = target_folder.Items
        messages.Sort("[ReceivedTime]", True)
        
        for msg in list(messages)[:6]: # on lui fait regarder les 6 derniers mail dans les "rappports du matin"
            try:
                sujet = msg.Subject.lower()
                
                # 1. On cherche d'abord les fichiers CSV par le nom des pièces jointes
                # En gros je lui demande de regarder dans les 6 derniers mail si une pièces jointe porte ce nom, si oui il la télécharge dans le fichier temp
                for att in msg.Attachments:
                    nom_att = att.FileName.lower()
                    
                    # Si c'est le fichier PRTG 
                    if "prtg" in nom_att and nom_att.endswith('.csv'):
                        att.SaveAsFile(os.path.join(FICHIER_TEMP, "prtg_sites.csv"))
                        print(f"-> CSV PRTG récupéré : {att.FileName}")
                    
                    # Si c'est le fichier METEO
                    elif ("meteo" in nom_att or "cd13" in nom_att) and nom_att.endswith('.csv'):
                        att.SaveAsFile(os.path.join(FICHIER_TEMP, "meteocd13.csv"))
                        print(f"-> CSV Météo récupéré : {att.FileName}")

                    elif "onepage" in nom_att and nom_att.endswith('.pdf'):
                        att.SaveAsFile(os.path.join(FICHIER_SORTIE, "Top_Applications_"+date_aujourdhui+".pdf"))

                # 2. je lui demande de regarder dans les mails si y a une image qui fini en .png, .jgp, ... 
                #Si y a une des images je lui demande de l'enregistrer dans le fichier temps et de l'appeler de manière incrémental car il s'agit d'un compteur 1 puis 2 puis 3 ....
                if any(word in sujet for word in ["check","image", "prtg", "météo","panorama"]):
                    compteur = 0
                    for att in msg.Attachments:
                        if att.FileName.lower().endswith(('.png', '.jpg', '.jpeg')):
                            nom_img = f"photo_{compteur}.jpg"
                            att.SaveAsFile(os.path.join(FICHIER_TEMP, nom_img))
                            compteur += 1
                    if compteur > 0:
                        print(f"-> {compteur} images récupérées dans le mail : {msg.Subject}") #la je fais une vérif je lui demande de me dire si compteur est supérieur a 0 de me dire la ou il a récupérer les mails et combien.

            except Exception as e:
                print(f"Erreur sur un mail : {e}") #Dans le cas ou y a une erreur je lui demande de me le dire
                continue
        return True
    except Exception as e:
        print(f"Erreur lors de la récupération : {e}") #Dans le cas ou y a une erreur je lui demande de me le dire
        return False

def generer_excel():
    print("Remplissage de l'Excel...")
    try:
        if not os.path.exists(TEMPLATE_CHEMIN): # Je lui demande de vérifier si le template est bien trouvable dans le chemin donnée
            print("ERREUR : Le fichier Template est introuvable !")
            return

        wb = load_workbook(TEMPLATE_CHEMIN) # Si il y a bien le template je le fais charger
        ws = wb.active

        #  rajout des images dans les casses du tableau E12 à E15
        for i in range(4):
            chemin_img = os.path.join(FICHIER_TEMP, f"photo_{i}.jpg")
            if os.path.exists(chemin_img):
                img = Image(chemin_img)
                img.width, img.height = 860, 280
                ws.add_image(img, f'E{12+i}')
                print(f"Image {i} insérée en E{12+i}")

        # Rajout des valeurs récupérer pour PRTG dans le tableau excel
        path_prtg = os.path.join(FICHIER_TEMP, "prtg_sites.csv")
        if os.path.exists(path_prtg):
            # Y avait un problème pendant l'encodage donc j'ai changer la méthode d'encodage par cp1252 pour les accents Windows (merci Gémini :) ) 
            df = pd.read_csv(path_prtg, sep=",", encoding='cp1252')
            ko = df[df['Statut'] == 'Erreur']
            ws['E9'] = ", ".join(ko[ko['Balises'] == 'sitec']['Groupe'].astype(str).tolist()) or "Ø"
            ws['E10'] = ", ".join(ko[ko['Balises'] == 'sited']['Groupe'].astype(str).tolist()) or "Ø"

        #  Rajout des valeur traité dans le fichier meteocd13.csv dans l'emplacement indiqué
        path_meteo = os.path.join(FICHIER_TEMP, "meteocd13.csv")
        if os.path.exists(path_meteo):
            dfm = pd.read_csv(path_meteo, sep=",", encoding='cp1252')
            # Filtrage des  Colonnes E 'active' et Col H 'down'
            ko_m = dfm[(dfm.iloc[:, 4] == 'active') & (dfm.iloc[:, 7].str.lower() == 'down')]
            ws['E8'] = ", ".join(ko_m.iloc[:, 2].astype(str).tolist()) or "Ø"
        
        chaine_bug = "274, 176, 687, 687, 687, 153"

        if ws['E8'].value == chaine_bug  :
            ws['E8'] = "Ø"
            print(f"remplacement des faux positif du dossier météoCD13")   


        # il save et il l'ouvre la
        wb.save(nouveau_fichier)
        print(f"TERMINÉ ! Checklist créée ici : {nouveau_fichier}")
        os.startfile(nouveau_fichier)

        #Dans le cas ou y a un problème avec la génération du excel
    except Exception as e:
        print(f"Erreur lors de la génération Excel : {e}")

if __name__ == "__main__":
    # Nettoyage du dossier temp avant de commencer
    for f in os.listdir(FICHIER_TEMP):
        try: os.remove(os.path.join(FICHIER_TEMP, f))
        except: pass
        
    if recuperer_mails_et_pieces():
        generer_excel()